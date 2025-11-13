# backend/logic.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import numpy as np
import re
import unicodedata
import plotly.graph_objects as go
from typing import Optional, Tuple, List, Dict, Any
import io
import json
import base64
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# --- RUTAS A ARCHIVOS FIJOS ---
BASE_DIR = os.path.dirname(__file__)
RUTA_DICC = os.path.join(BASE_DIR, "data", "Diccionario NFV_2.xlsx")
RUTA_CATALOGO = os.path.join(BASE_DIR, "data", "CID_Diferencias_catalogo_Vs_Antiago_250825.xlsx")

# ==========================================================
# SECCIÓN 1: LÓGICA DE LA MACRO (Leer TXT -> Excel)
# ==========================================================
# (Esta sección no cambia)
def importar_txt_a_excel(path_dir, nombre_sitio="default"):
    archivos = []
    for root, dirs, files in os.walk(path_dir):
        for file in files:
            if file.endswith('.txt'):
                full_path = os.path.join(root, file)
                sheet_name = os.path.splitext(file)[0][:31]
                archivos.append((full_path, sheet_name))
    wb = Workbook()
    wb.remove(wb.active)
    print(f"Encontrados {len(archivos)} archivos .txt para procesar.")
    for ruta, nombre_hoja in archivos:
        df = pd.read_csv(ruta, delimiter='|', encoding='utf-8',header=1)
        df = df.iloc[1:].reset_index(drop=True)
        df = df.drop(df.columns[0], axis=1)
        df = df.drop(df.columns[-1], axis=1)
        df = df.dropna(how='all')
        ws = wb.create_sheet(title=nombre_hoja)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        df2 = pd.read_csv(ruta, delimiter='|', usecols=[0], header=None, encoding='utf-8')
        df2.columns = [""]
        df2 = df2.dropna(how='all')
        df2 = df2.iloc[3:].reset_index(drop=True)
        for r in dataframe_to_rows(df2, index=False, header=True):
            ws.append(r)
    wb.create_sheet(title="Consolidado")
    ruta_salida = os.path.join(path_dir, f"Macro_Impresos_Generada.xlsx")
    wb.save(ruta_salida)
    return ruta_salida

def consolidar_datos(path_excel):
    """
    QUÉ HACE: Abre el Excel generado y combina todas las hojas
    en la hoja "Consolidado".
    """
    wb = load_workbook(path_excel)
    ws_consolidado = wb["Consolidado"]
    ws_consolidado.append(["Region", "Sitio", "ID", "Name", "OS-EXT-SRV-ATTR: Hypervisor Hostname",
                           "OS-EXT-SRV-ATTR: Instance Name", "OS-EXT-AZ: Availability Zone",
                           "flavor: Ram", "flavor: Vcpus", "flavor: Disk", "Status"])
    
    for sheet in wb.sheetnames:
        if sheet == "Consolidado": continue
        ws = wb[sheet]
        region = sheet[:2]
        sitio = sheet.split('_', 1)[1] if '_' in sheet else sheet
        sitio = sitio.split('-', 1)[0] if '-' in sitio else sheet
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): break
            fila = [region, sitio] + list(row[:10])
            data.append(fila)
        for row in data:
            ws_consolidado.append(row)
    wb.save(path_excel)

def crear_segunda_parte(path_excel):
    """
    QUÉ HACE: Lee la segunda parte de cada hoja (después de la
    fila vacía) y la consolida en una nueva hoja "2daParte".
    Formatea la AZ como 'AZ##'.
    """
    wb = load_workbook(path_excel)
    data = []
    for sheet in wb.sheetnames:
        if sheet in ["Consolidado", "2daParte"]: continue
        ws = wb[sheet]
        region = sheet[:2]
        sitio = sheet.split('_', 1)[1] if '_' in sheet else sheet
        sitio = sitio.split('-', 1)[0] if '-' in sitio else sheet
        comenzar = False
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if not any(row):
                comenzar = True
                continue
            if not comenzar: continue
            col_c = row[0]
            col_c = col_c.strip() if isinstance(col_c, str) else ''
            data.append([region, sitio, col_c])

    df = pd.DataFrame(data, columns=["Region", "Sitio", "Azone Hostname"])
    df["Azone"] = df["Azone Hostname"].apply(lambda x: x.split(" ")[0] if isinstance(x, str) else "")
    df["Hostname"] = df["Azone Hostname"].apply(lambda x: x.split(" ", 1)[1] if isinstance(x, str) and " " in x else x)
    
    def format_az_from_azone(x):
        if x in ["internal", "nova"]:
            return x
        az_str = str(x)
        
        match = re.search(r"(az\d{2})", az_str, re.IGNORECASE)
        if match:
            return match.group(1).upper()
        
        match_cee = re.search(r"cee(\d{2})", az_str, re.IGNORECASE)
        if match_cee:
            return "AZ" + match_cee.group(1)
        
        match_num = re.search(r"(\d{2})$", az_str)
        if match_num:
             return "AZ" + match_num.group(1)
            
        return "AZ_ND"

    df["AZ"] = df["Azone"].apply(format_az_from_azone)
    
    df["Host"] = df.apply(lambda row: row["Hostname"].split(".")[0] if isinstance(row["Hostname"], str) and "." in row["Hostname"] else row["Azone Hostname"], axis=1)
    df = df.drop(columns=["Azone Hostname"])
    
    ws_2da = wb.create_sheet(title="2daParte")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_2da.append(r)
    wb.save(path_excel)

def ejecutar_macro_completa(path_dir):
    """
    QUÉ HACE: Orquesta los 3 pasos anteriores.
    DEVUELVE: La ruta al excel final.
    """
    archivo_excel = importar_txt_a_excel(path_dir)
    consolidar_datos(archivo_excel)
    crear_segunda_parte(archivo_excel)
    print("✅ Consolidado completado. Archivo guardado en:", archivo_excel)
    return archivo_excel

# ==========================================================
# SECCIÓN 2: LÓGICA DE MAPEADO (VNF/VNFC)
# ==========================================================
# (Esta sección no cambia)
def _norm(s: str) -> str:
    if pd.isna(s): return ""
    s = str(s).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return " ".join(s.split())

def _best_match(texto: str, candidatos, whole_word: bool = False):
    textoN = _norm(texto)
    mejor, mejor_len = None, -1
    for cand in candidatos:
        cN = _norm(cand)
        if not cN: continue
        if whole_word:
            pat = rf'(?<![A-Z0-9]){re.escape(cN)}(?![A-Z0-9])'
            if re.search(pat, textoN):
                L = len(cN)
                if L > mejor_len:
                    mejor, mejor_len = cand, L
        else:
            if cN in textoN:
                L = len(cN)
                if L > mejor_len:
                    mejor, mejor_len = cand, L
    return mejor

def mapear_vnf_vnfc(df_names: pd.DataFrame, df_dicc: pd.DataFrame,
                    col_name="Name") -> pd.DataFrame:
    dicc = df_dicc.copy()
    necesarios = {"VNF", "VNFC", "VNF2", "VM Type", "Name contiene"}
    if not necesarios <= set(dicc.columns):
        raise ValueError(f"Faltan columnas en el diccionario: {necesarios - set(dicc.columns)}")

    firmas_por_vnf2 = {}
    for _, row in dicc.iterrows():
        vnf2 = row["VNF2"]
        if pd.isna(vnf2) or not str(vnf2).strip(): continue
        vnf2 = str(vnf2)
        firma = row.get("Name contiene", None)
        firmas_por_vnf2.setdefault(vnf2, set())
        if pd.notna(firma) and str(firma).strip():
            firmas_por_vnf2[vnf2].add(str(firma))
        firmas_por_vnf2[vnf2].add(vnf2)

    vmtypes_por_vnf2 = {}
    vnf_por_vnf2 = {}
    for _, row in dicc.iterrows():
        vnf2 = row["VNF2"]
        if pd.isna(vnf2) or not str(vnf2).strip(): continue
        vnf2 = str(vnf2)
        vnf = row["VNF"]
        if vnf_por_vnf2.get(vnf2) in (None, "") and pd.notna(vnf) and str(vnf).strip():
            vnf_por_vnf2[vnf2] = str(vnf)
        vm_t = row["VM Type"]
        vnfc = row["VNFC"]
        if pd.notna(vm_t) and str(vm_t).strip() and pd.notna(vnfc) and str(vnfc).strip():
            vmtypes_por_vnf2.setdefault(vnf2, [])
            vmtypes_por_vnf2[vnf2].append((str(vm_t), str(vnfc)))

    def resolver_name(name: str) -> Tuple[Optional[str], Optional[str]]:
        if pd.isna(name) or not str(name).strip():
            return (None, None)
        mejor_vnf2, mejor_len = None, -1
        for vnf2, firmas in firmas_por_vnf2.items():
            firma_match = _best_match(name, firmas, whole_word=False)
            if firma_match is not None:
                L = len(_norm(firma_match))
                if L > mejor_len:
                    mejor_vnf2, mejor_len = vnf2, L
        if not mejor_vnf2:
            return (None, None)
        vnf = vnf_por_vnf2.get(mejor_vnf2, None)
        pares_vmtype_vnfc = vmtypes_por_vnf2.get(mejor_vnf2, [])
        if not pares_vmtype_vnfc:
            return (vnf, None)
        mejor_vmtype, mejor_vnfc, mejor_len2 = None, None, -1
        for vm_t, vnfc_cand in pares_vmtype_vnfc:
            vm_match = _best_match(name, [vm_t], whole_word=True)
            if vm_match is not None:
                L2 = len(_norm(vm_match))
                if L2 > mejor_len2:
                    mejor_len2, mejor_vmtype, mejor_vnfc = L2, vm_t, vnfc_cand
        if mejor_vnfc is None:
            for vm_t, vnfc_cand in pares_vmtype_vnfc:
                vm_match = _best_match(name, [vm_t], whole_word=False)
                if vm_match is not None:
                    L2 = len(_norm(vm_match))
                    if L2 > mejor_len2:
                        mejor_len2, mejor_vmtype, mejor_vnfc = L2, vm_t, vnfc_cand
        return (vnf, mejor_vnfc)

    vnf_list, vnfc_list = [], []
    for name in df_names[col_name]:
        v, c = resolver_name(name)
        vnf_list.append(v)
        vnfc_list.append(c)
    return pd.DataFrame({"VNF": vnf_list, "VNFC": vnfc_list})

def completar_excepciones(df_names: pd.DataFrame,
                          mapeo: pd.DataFrame,
                          df_exc: pd.DataFrame,
                          col_name="Name") -> pd.DataFrame:
    out = mapeo.copy()
    if not {"VNF", "TIPO VM"} <= set(df_exc.columns):
        raise ValueError("La hoja 'Excepciones' debe tener columnas: 'VNF' y 'TIPO VM'.")
    lista_vnf_exc = [str(v) for v in df_exc["VNF"].dropna().astype(str) if str(v).strip()]

    for i, (vnf_val, vnfc_val) in enumerate(zip(out["VNF"], out["VNFC"])):
        if pd.isna(vnf_val) and pd.isna(vnfc_val):
            name = df_names.iloc[i][col_name]
            if pd.isna(name) or not str(name).strip(): continue
            vnf_match = _best_match(name, lista_vnf_exc, whole_word=True)
            if vnf_match is None:
                vnf_match = _best_match(name, lista_vnf_exc, whole_word=False)
            if vnf_match is not None:
                fila = df_exc[df_exc["VNF"].astype(str) == str(vnf_match)]
                tipo_vm = None
                if not fila.empty:
                    tipo_vm_series = fila["TIPO VM"].dropna().astype(str)
                    if not tipo_vm_series.empty:
                        tipo_vm = tipo_vm_series.iloc[0]
                out.at[i, "VNF"] = str(vnf_match)
                out.at[i, "VNFC"] = tipo_vm if tipo_vm is not None else None
    return out

# ==========================================================
# SECCIÓN 3: LÓGICA DE DIBUJADO (Plotly)
# ==========================================================
# (Esta sección no cambia)
def f_tetris_plot(tetris_piezas: pd.DataFrame, tamaño_chip: int, sin_uso: int, site_name: str, cee_tag: str):
    df = tetris_piezas.copy()
    df['AZ'] = df['AZ'].fillna('AZ_ND')
    df['Host'] = df['Host'].fillna('')
    df['CEE'] = df['CEE'].fillna('') 
    
    df['host_chip'] = df['CEE'].astype(str) + ' - ' + df['AZ'] + ' - ' + df['Host'].astype(str)
    
    ordered_hosts = (
        df[['AZ', 'Host', 'host_chip']]
        .drop_duplicates()
        .sort_values(by=['AZ', 'Host'])['host_chip']
        .tolist()[::-1]
    )
    fig = go.Figure()
    
    title_text = f'TETRIS actual - {site_name} - {cee_tag}'
    
    for _, row in df.iterrows():
        if row["VM_Individual_Count"] > 0:
            pieza_str = f'Pieza: {row["VM_Individual_Count"]} de {row["VM_Total_Count"]}'
        else:
            pieza_str = ''

        hover_template = (
            f'<b>{row["VM"]}</b><br>'
            f'Longitud: {row["Length"]}<br>'
            f'Anti-Afinidad: {int(row["Anti-Afinidad"])}<br>'
            f'{pieza_str}<extra></extra>'
        ) if pd.notnull(row['VM']) else '<extra></extra>'

        fig.add_trace(go.Bar(
            x=[row['Length']],
            y=[row['host_chip']],
            text=row['VM'],
            base=row['Start'],
            name=row['VM'],
            orientation='h',
            marker=dict(color=row['Color']),
            hovertemplate=hover_template,
            showlegend=False
        ))
    fig.update_layout(
        barmode='stack',
        title=title_text,
        xaxis=dict(
            tickmode='linear',
            dtick=1,
            range=[0, 2 * tamaño_chip + 2 * sin_uso],
            showgrid=True
        ),
        yaxis=dict(
            title='Host',
            categoryorder='array',
            categoryarray=ordered_hosts
        ),
        height = 2000,
        width=1500,
        showlegend=False
    )
    fig.update_traces(
        textposition='inside',
        insidetextanchor='middle',
        textfont=dict(color='black', size=10),
        marker=dict(line=dict(width=1.5, color='black'))
    )
    return fig

def build_tetris_for_cee(data, zonas, cee_value, ancho_cee=42):
    tetris = []
    
    if cee_value is None:
        data_cee_full = data[data['CEE'].isna()]
    else:
        data_cee_full = data[data['CEE'].astype(str) == str(cee_value)]

    zonas_en_cee = sorted(data_cee_full["OS-EXT-AZ_Formatted"].dropna().unique())

    for zona in zonas_en_cee: 
        data_cee = data_cee_full[data_cee_full["OS-EXT-AZ_Formatted"] == zona]
        if data_cee.empty: continue
        
        short_zona = zona 
       
        zona_larga_series = data_cee['OS-EXT-AZ: Availability Zone'].dropna()
        zona_larga = zona_larga_series.iloc[0] if not zona_larga_series.empty else short_zona
        
        tetris.append({'AZ': short_zona, 'Host': ' ', 'Chip': ' ', 'VM': zona_larga,
                       'Start': 0, 'Length': ancho_cee, 'Color': '#696969',
                       'Inst.': ' ', 'Anti-Afinidad': 0, 'Requerimiento': ' ', 'HOST': ' '})
        hosts = data_cee['OS-EXT-SRV-ATTR: Hypervisor Hostname'].unique()
        for host_name in hosts:
            data_host = data_cee[data_cee['OS-EXT-SRV-ATTR: Hypervisor Hostname'] == host_name] \
                .sort_values(by='flavor: Vcpus', ascending=False)
            chip1 = 0
            chip2 = 20
            cee_tag = f'CEE{cee_value}' if cee_value is not None else 'CEE_NA'
            
            tetris.append({'AZ': short_zona, 'Host': host_name, 'Chip': ' ', 'VM': ' ',
                           'Start': 17, 'Length': 3, 'Color': '#000000',
                           'Inst.': ' ', 'Anti-Afinidad': 0, 'Requerimiento': ' ', 'HOST': cee_tag})

            tetris.append({'AZ': short_zona, 'Host': host_name, 'Chip': ' ', 'VM': ' ',
                           'Start': 37, 'Length': 3, 'Color': '#000000',
                           'Inst.': ' ', 'Anti-Afinidad': 0, 'Requerimiento': ' ', 'HOST': cee_tag})

            for _, row in data_host.iterrows():
                tamaño = row['flavor: Vcpus'] / 2
                if chip1 + tamaño <= 17:
                    chip = 'Chip 1'
                    start_rel = chip1
                    chip1 += tamaño
                else:
                    chip = 'Chip 2'
                    start_rel = chip2
                    chip2 += tamaño

                tetris.append({'AZ': short_zona, 'Host': host_name, 'Chip': chip,
                               'VM': row['Name'], 'Start': start_rel, 'Length': tamaño,
                               'Color': row['Color'], 'Inst.': ' ', 'Anti-Afinidad': 0,
                               'Requerimiento': ' ', 'HOST': cee_tag})
    
    tetris_df = pd.DataFrame(tetris)
    if tetris_df.empty: return tetris_df
    if 'HOST' not in tetris_df.columns: tetris_df['HOST'] = ' '
    
    tetris_df = tetris_df.sort_values(by=['AZ', 'HOST', 'Host', 'Chip', 'Start']).reset_index(drop=True)

    vm_rows = tetris_df['VM'].str.strip().fillna('').ne('') & tetris_df['Length'].gt(0)
    vm_groups = tetris_df[vm_rows].groupby('VM')
    tetris_df.loc[vm_rows, 'VM_Total_Count'] = vm_groups['VM'].transform('count')
    tetris_df.loc[vm_rows, 'VM_Individual_Count'] = vm_groups.cumcount() + 1
    tetris_df['VM_Total_Count'] = tetris_df['VM_Total_Count'].fillna(0).astype(int)
    tetris_df['VM_Individual_Count'] = tetris_df['VM_Individual_Count'].fillna(0).astype(int)
    
    tetris_df['Numero'] = tetris_df.apply(
        lambda row: f"{row['VM_Individual_Count']} de {row['VM_Total_Count']}" if row['VM_Individual_Count'] > 0 else '',
        axis=1
    )
    
    tetris_df['CEE'] = tetris_df['HOST']
    unique_hosts = tetris_df['Host'].unique()
    real_hosts = [h for h in unique_hosts if h.strip() != '']
    host_map = {host_name: i + 1 for i, host_name in enumerate(real_hosts)}
    tetris_df['HOST'] = tetris_df['Host'].map(host_map)
    tetris_df['HOST'] = tetris_df['HOST'].astype('object').fillna(' ') 

    host_relativos = []
    for zona, grupo in tetris_df.groupby("AZ", sort=False):
        unique_hosts_zona = grupo['Host'].unique()
        real_hosts_zona = [h for h in unique_hosts_zona if h.strip() != '']
        host_map_zona = {host_name: i + 1 for i, host_name in enumerate(real_hosts_zona)}
        host_relativos.extend(grupo['Host'].map(host_map_zona))
        
    tetris_df['HOST_RELATIVO'] = host_relativos
  
    return tetris_df

# ==========================================================
# SECCIÓN 4: FUNCIÓN "CEREBRO" PRINCIPAL
# ==========================================================
def generar_reportes_tetris(temp_dir_path: str, site_name: str) -> List[dict]:
    
    ruta_macro = ejecutar_macro_completa(temp_dir_path)
    
    print("Cargando diccionarios...")
    df_dicc = pd.read_excel(RUTA_DICC, sheet_name="Diccionario")
    df_exc  = pd.read_excel(RUTA_DICC, sheet_name="Excepciones")

    print("Cargando datos consolidados...")
    data = pd.read_excel(ruta_macro, sheet_name="Consolidado") 
    
    print("Enriqueciendo datos de AZ...")
    df_2daparte = pd.read_excel(ruta_macro, sheet_name="2daParte")
    data['Host_short'] = data['OS-EXT-SRV-ATTR: Hypervisor Hostname'].astype(str).str.split('.').str[0]
    df_2daparte_clean = df_2daparte[['Host', 'AZ']].drop_duplicates().dropna(subset=['Host'])
    host_az_map = df_2daparte_clean.set_index('Host')['AZ'].to_dict()
    
    def fill_missing_az(row):
        az = row['OS-EXT-AZ: Availability Zone']
        if pd.isna(az) or str(az).strip() == '':
            return host_az_map.get(row['Host_short'], az)
        return az

    data['OS-EXT-AZ: Availability Zone'] = data.apply(fill_missing_az, axis=1)
    
    def format_az_column(az):
        if pd.isna(az) or str(az).strip() == '':
            return "AZ_ND"
        az_str = str(az)
        if az_str in ["internal", "nova"]:
            return az_str
        
        match = re.search(r"(az\d{2})", az_str, re.IGNORECASE)
        if match:
            return match.group(1).upper()
        
        match_cee = re.search(r"cee(\d{2})", az_str, re.IGNORECASE)
        if match_cee:
            return "AZ" + match_cee.group(1)
        
        match_num = re.search(r"(\d{2})$", az_str)
        if match_num:
             return "AZ" + match_num.group(1)

        return "AZ_ND" 

    data["OS-EXT-AZ_Formatted"] = data['OS-EXT-AZ: Availability Zone'].apply(format_az_column)
    
    zonas = sorted(data["OS-EXT-AZ: Availability Zone"].dropna().unique())

    print("Mapeando VNF/VNFC...")
    mapeo_inicial = mapear_vnf_vnfc(data, df_dicc, col_name="Name")
    mapeo_final = completar_excepciones(data, mapeo_inicial, df_exc, col_name="Name")
    v = mapeo_final["VNF"].astype(str).str.strip().replace({"<NA>": np.nan, "nan": np.nan, "": np.nan})
    c = mapeo_final["VNFC"].astype(str).str.strip().replace({"<NA>": np.nan, "nan": np.nan, "": np.nan})
    data["Name"] = np.where(~v.isna() & ~c.isna(), v + " " + c,
                        np.where(~v.isna(), v, data["Name"]))

    print("Asignando colores...")
    base = pd.read_excel(RUTA_CATALOGO, header=1)
    base["VNF"] = base["VNF"].astype(str)
    data["Name"] = data["Name"].astype(str)
    def obtener_color(name):
        for _, fila in base.iterrows():
            if fila["VNF"] in name:
                return fila["Color"]
        return None
    data["Color"] = data["Name"].apply(obtener_color)
    data["Color"].fillna("#808080", inplace=True) 

    print("Extrayendo CEEs...")
    data["CEE"] = data["OS-EXT-AZ: Availability Zone"].astype(str).apply(
        lambda x: re.search(r"vapp(\d+)", x).group(1) if re.search(r"vapp(\d+)", x) else None
    )

    print("Generando gráficos...")
    cee_vals = pd.to_numeric(data["CEE"], errors="coerce")
    unique_cees = sorted([int(x) for x in cee_vals.dropna().unique()])
    if data["CEE"].isna().any():
        unique_cees.append(None)
    
    tamaño_chip = 17
    sin_uso = 3
    
    html_outputs = [] 
    for cee in unique_cees:
        cee_tag = f"CEE{cee}" if cee is not None else "CEE_NA"
        print(f"Procesando {cee_tag}...")

        tetris_cee = build_tetris_for_cee(data, None, cee, ancho_cee=42)
        if tetris_cee.empty:
            print(f"No hay piezas para {cee_tag}, se omite.")
            continue
        
        # Lógica  de reemplazar ' ' por 'INFRA'
        tetris_cee.loc[tetris_cee["VM"] == " ", "VM"] = "INFRA"

        df_plotly = tetris_cee.copy() # Usar una copia para Plotly
        df_plotly['AZ'] = df_plotly['AZ'].fillna('AZ_ND')
        df_plotly['Host'] = df_plotly['Host'].fillna('')
        df_plotly['CEE'] = df_plotly['CEE'].fillna('')
        df_plotly['host_chip'] = df_plotly['CEE'].astype(str) + ' - ' + df_plotly['AZ'] + ' - ' + df_plotly['Host'].astype(str)
        tetris_cee['host_chip'] = df_plotly['host_chip'] # Asegurar que esta columna existe para CSV
        base_filename = f"{site_name}_tetris_{cee_tag}"
        
        # ==================================================
        # CAMBIO CRÍTICO: CALCULAR Y GUARDAR ESTADÍSTICAS
        # ==================================================
        try:
            stats_cee = calculate_statistics(tetris_cee)
            stats_json = json.dumps(stats_cee)
            
            # Agregamos el JSON a la lista de outputs para que el frontend lo reciba
            html_outputs.append({
                "filename": f"{site_name}_tetris_{cee_tag}_stats.json", 
                "content": stats_json
            })
            print(f"📊 Estadísticas generadas para {cee_tag}")
        except Exception as e:
            print(f"⚠️ Error calculando estadísticas para {cee_tag}: {e}")
        # ==================================================

        # --- 1. Generar HTML ---
        fig = f_tetris_plot(df_plotly, tamaño_chip, sin_uso, site_name, cee_tag)
        html_content = fig.to_html(full_html=True, include_plotlyjs=True)
        html_outputs.append({
            "filename": f"{base_filename}.html",
            "content": html_content
        })

        # --- 2. Generar CSV ---
        print(f"Generando CSV para {cee_tag}...")
        columnas_map = {
            'AZ': 'AZ', 'Host': 'HOST_RELATIVO', 'Chip': 'Chip', 'VM': 'VM',
            'Start': 'Start', 'Length': 'Length', 'Color': 'Color', 'Inst.': 'Inst.',
            'Anti-Afinidad': 'Anti-Afinidad', 'HOST': 'HOST', 'Numero': 'Numero',
            'host_chip': 'host_chip'
        }
        df_csv_final = pd.DataFrame()
        for col_deseada, col_original in columnas_map.items():
            if col_original in tetris_cee.columns:
                df_csv_final[col_deseada] = tetris_cee[col_original]
            else:
                df_csv_final[col_deseada] = np.nan 
        csv_content = df_csv_final.to_csv(index=False, encoding='utf-8-sig')
        html_outputs.append({
            "filename": f"{base_filename}.csv",
            "content": csv_content
        })

        # --- 3. Generar Excel ---
        print(f"Generando Excel para {cee_tag}...")
        try:
            # 1. Crear el diccionario de COLORES 
            diccionario_colores = generar_diccionario_colores(tetris_cee) 
            # 2. Crear la MATRIZ, pero almacenando un TUPLE (VM, Length) 
            matriz_excel = generar_matriz_con_length(tetris_cee)
            # 3. Definir columnas 
            columnas_excel = definir_columnas(col_info)
            # 4. Obtener DF final, aplicando el REORDENAMIENTO de columnas
            df_excel_final = obtener_df_final(matriz_excel, columnas_excel)
            
            # 5. Generar bytes, adaptado para leer el TUPLE (VM, Length)
            excel_bytes = generar_excel_bytes(site_name, df_excel_final, diccionario_colores)
         
            excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
            
            html_outputs.append({
                "filename": f"{base_filename}.xlsx",
                "content": excel_base64
            })
            print(f"✅ Excel generado para {cee_tag}")

        except Exception as e:
            print(f"❌ Error generando Excel para {cee_tag}: {e}")

        # 4. NUEVO: Agregar Estadísticas al output
        # Lo pasaremos como un objeto JSON separado o embebido si el front lo soporta.
        # Dado que app.js agrupa por nombre base, podemos enviar un .json

        stats_json = json.dumps(stats_cee)
        html_outputs.append({
            "filename": f"{base_filename}_stats.json", # App.js detectará esto
            "content": stats_json # Enviar string JSON
        })

    print("✅ Proceso completado.")
    return html_outputs

# ==========================================================
# SECCIÓN 5: LÓGICA DE EXCEL 
# ==========================================================

# --- Constantes ---
NUMERO_CHIPS = 2
NUMERO_CORES = 20
col_info = ['# S. TOTAL', '# S. ZONA', 'SERVIDOR']

def generar_diccionario_colores(df_solucion: pd.DataFrame) -> Dict[str, str]:
    """
    Crea un diccionario de VM -> Color.
    """
    df_diccionario = df_solucion[['VM', 'Color']]
    df_unique = df_diccionario.drop_duplicates(subset="VM")
    diccionario_colores = {fila.VM: fila.Color for fila in df_unique.itertuples(index=False)}
    diccionario_colores['INFRA'] = '#FBE2D5' 
    
    # Manejar VMs sin color
    for vm, color in diccionario_colores.items():
        if pd.isna(color):
            diccionario_colores[vm] = "#808080" # Gris por defecto
            
    return diccionario_colores

# --- Genera la matriz con (VM, Length) en las celdas ---
def generar_matriz_con_length(df_solucion: pd.DataFrame) -> List[List[Any]]:
    """
    Convierte el dataframe 'tetris_cee' en una matriz ancha para Excel.
    Cada celda que representa una VM ahora contiene un TUPLE
    (VM_Nombre, Length).
    """
    df_solucion['HOST_RELATIVO'] = df_solucion['HOST_RELATIVO'].fillna(0)

    servidores = df_solucion['HOST'].unique()
    servidores_reales = [h for h in servidores if h != ' ']
    if not servidores_reales:
        print("Advertencia: No se encontraron servidores reales para la matriz de Excel.")
        return []
    
    # Crear un mapa de Host (nombre largo) a su # S. TOTAL (HOST)
    host_to_num = df_solucion.drop_duplicates(subset='Host').set_index('Host')['HOST'].to_dict()

    # Mapeo de HOST (número total) a índice de fila en la matriz (0-based)
    servidores_reales_sorted = sorted([h for h in host_to_num.values() if h != ' '])
    host_num_to_matriz_idx = {host_num: i for i, host_num in enumerate(servidores_reales_sorted)}
    
    matriz = [[np.nan for _ in range((NUMERO_CORES*NUMERO_CHIPS) + len(col_info) + 1)] for _ in range(len(servidores_reales))]

    for fila in df_solucion.itertuples(index=False):
        host_num = host_to_num.get(fila.Host)
        if host_num is None or host_num == ' ': continue 
        
        idx_fila = host_num_to_matriz_idx.get(host_num)
        if idx_fila is None: continue # Host no está en el mapa (ej. cabecera AZ)
        
        try:
            idx_col = int(fila.Start)
        except ValueError:
            continue
        # Almacenar una tupla (Nombre, Longitud)
        matriz[idx_fila][idx_col] = (fila.VM, fila.Length)
            
        matriz[idx_fila][-4] = fila.AZ if pd.notna(fila.AZ) else "AZ_ND"
        matriz[idx_fila][-3] = fila.HOST 
        matriz[idx_fila][-2] = int(fila.HOST_RELATIVO)
        matriz[idx_fila][-1] = fila.Host 
    
    matriz = [fila for fila in matriz if not all(pd.isna(c) for c in fila)]
    return matriz

def definir_columnas(col_info):
    """Define las columnas para el DataFrame ancho."""
    # Define el orden base, la reorganización se hace en obtener_df_final
    return [f'chip_{i}_{j}' for i in range(1, NUMERO_CHIPS + 1) for j in range(1, NUMERO_CORES + 1)] + ['AZ'] + col_info

def obtener_df_final(matriz: List[List[Any]], columnas: List[str]) -> pd.DataFrame:
    """Crea el DataFrame final y añade las filas de separación por AZ."""
    df = pd.DataFrame(matriz, columns=columnas)
    
    if 'SERVIDOR' in df.columns:
        df['SERVIDOR'] = df['SERVIDOR'].replace(' ', 'INFRA')

    if not df.empty:
        # Lógica para insertar separadores
        df_az_str = df["AZ"].astype(str) # Convertir a string para comparación segura
        cambios = df_az_str != df_az_str.shift()
        filas_cambio = df.index[cambios].tolist()[1:]
        for idx in reversed(filas_cambio):
            df = pd.concat([
                df.iloc[:idx],
                pd.DataFrame([{}], columns=df.columns), # Fila vacía
                df.iloc[idx:]
            ], ignore_index=True)
    # reordenar columnas
    cols = df.columns
    try:
        # Orden: [INFO] + [CHIP 1] + [AZ] + [CHIP 2]
        df = df[cols[-len(col_info):].tolist() + cols[:NUMERO_CORES].tolist() + [cols[-(len(col_info)+1)]] + cols[NUMERO_CORES:-(len(col_info)+1)].tolist()]
    except Exception as e:
        print(f"Error reordenando columnas (se omite): {e}")
   
    return df

def generar_excel_bytes(sitio: str, df: pd.DataFrame, diccionario_colores: Dict[str, str]) -> bytes:
    """
    Toma el DataFrame final, lo escribe en un Workbook de openpyxl,
    aplica todo el estilo de la Celda 18, y devuelve los bytes
    del archivo .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tetris"

    # 1. Escribir el DataFrame en la hoja
    # Itera el DF para obtener los valores originales (incluyendo los tuples)
    # Escribe solo el nombre de la VM en la celda
    
    # Escribir cabeceras
    for c_idx, col_name in enumerate(df.columns.tolist(), 1):
        ws.cell(row=2, column=c_idx, value=col_name)

    # Escribir datos
    for r_idx, row_data in enumerate(df.itertuples(index=False), 3): # Empezar en fila 3
        for c_idx, value in enumerate(row_data, 1):
            if isinstance(value, tuple) and len(value) == 2:
                ws.cell(row=r_idx, column=c_idx, value=value[0]) # Escribir solo VM
            else:
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 2. Insertar fila de título 
    filas_recorridas = 3 # 1 (título) + 1 (header) + 1 (excel 1-based)
    columnas_recorridas = 1 # Para el índice de 1 de openpyxl
    
    ws.cell(row=1, column=1, value=f"TETRIS ACTUAL {sitio}")

    # 3. Definir estilos
    fuente = Font(name="Aptos Narrow", size=11)
    alineacion = Alignment(horizontal="center", vertical="center")
    borde = Side(style="thin", color="000000")
    borde_completo = Border(left=borde, right=borde, top=borde, bottom=borde)

    # 4. Aplicar estilos generales
    fila_inicio_total = 1
    fila_fin_total = df.shape[0] + 2 # +2 por título y header
    columna_inicio_total = 1
    columna_fin_total = df.shape[1]

    for fila in range(fila_inicio_total, fila_fin_total + 1):
        for col in range(columna_inicio_total, columna_fin_total + 1):
            celda = ws.cell(row=fila, column=col)
            celda.font = fuente
            celda.alignment = alineacion
            celda.border = borde_completo

    # 5. Combinar celdas y colorear piezas
    cols_chip = [c for c in df.columns if c.startswith("chip_")]
    
    # Iterar sobre el DataFrame original (que tiene los tuples)
    for idx_df, row_data in enumerate(df.itertuples(index=False)):
        fila_excel = idx_df + filas_recorridas # Mapear al índice 1-based de Excel

        for col_name in cols_chip:
            if col_name not in df.columns: continue
            
            col_idx_df = df.columns.get_loc(col_name) # 0-based index en DF
            col_excel = col_idx_df + columnas_recorridas # 1-based index en Excel
            
            value_in_df = row_data[col_idx_df] # Obtener el valor (puede ser tuple)
            
            if pd.isna(value_in_df):
                continue
            
            vm_name, length, color = None, 0, None

            if isinstance(value_in_df, tuple) and len(value_in_df) == 2:
                vm_name, length = value_in_df
                color = diccionario_colores.get(vm_name)
            elif isinstance(value_in_df, str) and value_in_df == "INFRA":
                vm_name = "INFRA"
                length = 3 # Longitud fija para INFRA
                color = diccionario_colores.get("INFRA")
            else:
                continue # No es una pieza a pintar

            if length <= 0 or color is None:
                continue
                
            length = int(length)
            color_hex = str(color)[1:] # Quitar '#'

            fila_inicio_merge = fila_excel
            fila_fin_merge = fila_excel
            col_inicio_merge = col_excel
            col_fin_merge = col_excel + length - 1
            
            # Asegurarse de no combinar más allá del límite de columnas
            if col_fin_merge > columna_fin_total:
                col_fin_merge = columna_fin_total
            
            if col_inicio_merge <= col_fin_merge:
                try:
                    ws.merge_cells(start_row=fila_inicio_merge, start_column=col_inicio_merge, end_row=fila_fin_merge, end_column=col_fin_merge)
                    relleno = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    ws.cell(row=fila_inicio_merge, column=col_inicio_merge).fill = relleno
                except Exception as merge_err:
                    print(f"Error al fusionar celdas para {vm_name} en ({fila_inicio_merge},{col_inicio_merge})-({fila_fin_merge},{col_fin_merge}): {merge_err}")

    # 6. Colorear Headers (Fila 2) 
    fila_headers = 2
    color_header = 'FBE2D5' 
    relleno_header = PatternFill(start_color=color_header, end_color=color_header, fill_type="solid")
    for col in range(columna_inicio_total, columna_fin_total + 1):
        celda = ws.cell(row=fila_headers, column=col)
        celda.fill = relleno_header

        texto = celda.value
        if isinstance(texto, str) and "chip_" in texto:
            try:
                celda.value = int(texto.split("_")[-1])
            except:
                pass 
            
    # 7. Colorear columnas de info (# S. TOTAL y # S. ZONA)
    filas_datos_inicio = 3
    filas_datos_fin = df.shape[0] + 2 
    
    # Colorear # S. TOTAL y # S. ZONA (Columnas 1 y 2)
    for fila in range(filas_datos_inicio, filas_datos_fin + 1):
        for col in range(1, len(col_info)): # Col 1 y 2
            ws.cell(row=fila, column=col).fill = relleno_header
                
    # 8. Poner en negrita la columna SERVIDOR 
    try:
        col_servidor_idx_excel = df.columns.get_loc('SERVIDOR') + columnas_recorridas
        fuente_negrita = Font(name="Aptos Narrow", size=11, bold=True)
        for fila in range(filas_datos_inicio, filas_datos_fin + 1):
             ws.cell(row=fila, column=col_servidor_idx_excel).font = fuente_negrita
    except KeyError:
        print("Advertencia: No se encontró la columna 'SERVIDOR' para negrita.")

    # 9. Colorear filas vacías (separadores de AZ)
    indices_nulos = df.index[df.isnull().all(axis=1)].tolist()
    color_separador = "000000" # Negro
    relleno_separador = PatternFill(start_color=color_separador, end_color=color_separador, fill_type="solid")
    
    for idx_nulo in indices_nulos:
        fila_excel = idx_nulo + filas_recorridas 
        ws.merge_cells(start_row=fila_excel, start_column=columna_inicio_total, end_row=fila_excel, end_column=columna_fin_total)
        celda_merge = ws.cell(row=fila_excel, column=columna_inicio_total)
        celda_merge.fill = relleno_separador
        celda_merge.value = "INFRA" 
        
    # 10. Estilo del Título (Fila 1) 
    fuente_titulo = Font(name="Aptos Narrow", size=16, bold=True, color="FF0000")
    alineacion_titulo = Alignment(horizontal="center", vertical="center")
    color_titulo = "CAEDFB" # Azul claro
    relleno_titulo = PatternFill(start_color=color_titulo, end_color=color_titulo, fill_type="solid")
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columna_fin_total)
    celda_titulo = ws.cell(row=1, column=1)
    celda_titulo.font = fuente_titulo
    celda_titulo.alignment = alineacion_titulo
    celda_titulo.fill = relleno_titulo
    
    # 11. Estilizar y fusionar columna AZ (vertical)
    try:
        col_az_idx_excel = df.columns.get_loc('AZ') + columnas_recorridas
        fuente_az = Font(name="Aptos Narrow", size=14, bold=True)
        alineacion_az = Alignment(horizontal="center", vertical="center", textRotation=90)
        color_az = "B5E6A2" # Verde 
        relleno_az = PatternFill(start_color=color_az, end_color=color_az, fill_type="solid")
        
        fila_inicio_az_merge = 3 # Datos empiezan en fila 3
        df_az_clean = df.copy()
        df_az_clean['AZ_str'] = df_az_clean['AZ'].apply(lambda x: str(x) if not pd.isna(x) else 'NAN')
        
        az_blocks = []
        current_az = None
        start_row_block_df = -1

        for idx_df, az_val in enumerate(df_az_clean['AZ_str']):
            if az_val == 'NAN': 
                if current_az is not None:
                    az_blocks.append((current_az, start_row_block_df, idx_df - 1))
                    current_az = None
                continue
            
            if az_val != current_az:
                if current_az is not None:
                    az_blocks.append((current_az, start_row_block_df, idx_df - 1))
                current_az = az_val
                start_row_block_df = idx_df
        
        if current_az is not None:
            az_blocks.append((current_az, start_row_block_df, len(df_az_clean) - 1))

        for az_val, start_df_idx, end_df_idx in az_blocks:
            fila_inicio_excel = start_df_idx + filas_recorridas 
            fila_fin_excel = end_df_idx + filas_recorridas
            
            for fila_num in range(fila_inicio_excel, fila_fin_excel + 1):
                celda = ws.cell(row=fila_num, column=col_az_idx_excel)
                celda.font = fuente_az
                celda.alignment = alineacion_az
                celda.fill = relleno_az
            
            if fila_inicio_excel <= fila_fin_excel:
                 ws.merge_cells(start_row=fila_inicio_excel, start_column=col_az_idx_excel, 
                                 end_row=fila_fin_excel, end_column=col_az_idx_excel)
            
    except Exception as e:
        print(f"Error estilizando la columna AZ vertical: {e}")

    # 12. Guardar en un stream de bytes en lugar de un archivo
    with io.BytesIO() as bio:
        wb.save(bio)
        return bio.getvalue()
    
# --- AGREGAR ESTA FUNCIÓN AL FINAL O EN UNA SECCIÓN DE UTILIDADES ---
# --- AGREGAR EN backend/logic.py ---

def calculate_statistics(df: pd.DataFrame) -> Dict:
    HOST_CAPACITY = 34
    CHIP_CAPACITY = 17
    
    # Filtros básicos: Ignorar INFRA y filas vacías
    mask_reales = (df['VM'] != 'INFRA') & (df['VM'].str.strip() != '') & (df['Length'] > 0)
    df_reales = df[mask_reales].copy()
    
    # 1. Hosts reales
    # Convertimos a string para evitar errores si hay mezclas de tipos
    hosts_unicos = df[df['HOST'].astype(str).str.strip() != '']['HOST'].unique()
    used_hosts = len(hosts_unicos)
    
    # 2. Capacidad y Uso
    total_capacity = used_hosts * HOST_CAPACITY
    total_used = int(df_reales['Length'].sum()) if not df_reales.empty else 0
    
    # Evitar división por cero
    util = (total_used / total_capacity) if total_capacity > 0 else 0.0
    
    # --- CRÍTICO: n_vms necesario para app.js ---
    n_vms = len(df_reales)

    # 3. Estadísticas por AZ
    per_az = {}
    az_values = [] # Lista para app.js
    
    if used_hosts > 0 and 'AZ' in df.columns:
        for az, df_zona in df.groupby(by='AZ'):
            if str(az) == 'AZ_ND': continue
            az_values.append(str(az))
            
            # Hosts en esta AZ
            hosts_en_az = df_zona[df_zona['HOST'].astype(str).str.strip() != '']['HOST'].unique()
            n_hosts_az = len(hosts_en_az)
            
            # Uso en esta AZ
            mask_az = (df_zona['VM'] != 'INFRA') & (df_zona['VM'].str.strip() != '')
            used_in_az = int(df_zona.loc[mask_az, 'Length'].sum())
            
            cap_az = n_hosts_az * HOST_CAPACITY
            util_az = (used_in_az / cap_az) if cap_az > 0 else 0.0
            
            per_az[str(az)] = {
                "hosts": int(n_hosts_az),
                "used": used_in_az,
                "capacity": int(cap_az),
                "utilization": util_az
            }
    
    az_values.sort()

    # 4. Chips y Huecos
    chips_used = 0
    holes_hist = {}
    if used_hosts > 0 and not df_reales.empty:
        # Agrupar por HOST y Chip para unicidad
        for _, df_chip in df_reales.groupby(by=['HOST', 'Chip']):
            ch_used = df_chip["Length"].sum()
            slack = max(0, int(CHIP_CAPACITY - ch_used))
            holes_hist[str(slack)] = holes_hist.get(str(slack), 0) + 1
            chips_used += 1

    # 5. Utilización por Host
    host_utils = []
    if used_hosts > 0 and not df_reales.empty:
        for _, df_host in df_reales.groupby(by='HOST'):
            u = df_host["Length"].sum() / HOST_CAPACITY
            host_utils.append(u)
            
    # Valores seguros para evitar NaN en JSON (que rompe el JS)
    avg_util = sum(host_utils)/len(host_utils) if host_utils else 0.0
    max_util = max(host_utils) if host_utils else 0.0
    min_util = min(host_utils) if host_utils else 0.0

    stats = {
        "status": "REAL_LOG",
        "n_vms": n_vms,
        "az_values": az_values, 
        "hosts": used_hosts,
        "total_used": total_used,
        "total_capacity": total_capacity,
        "utilization": util,
        "empty_pct": 1 - util,
        "chips_used": chips_used,
        "holes_histogram": holes_hist,
        "host_utilization": {
            "avg": avg_util,
            "max": max_util,
            "min": min_util,
        },
        "per_az": per_az
    }
    return stats